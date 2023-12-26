# import tkinter as tk
# from tkinter import filedialog, Listbox, scrolledtext
# import openpyxl

# def load_excel_data():
#     file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
#     if file_path:
#         workbook = openpyxl.load_workbook(file_path)
#         sheet = workbook.active

#         listbox_items.clear()
#         excel_data.clear()
#         for row in sheet.iter_rows(min_row=2):  # Assuming the first row is the header
#             e_column_value = row[4].value if row[4].value is not None else ""
#             d_column_value = row[3].value if row[3].value is not None else ""
#             listbox_items.append(e_column_value)
#             excel_data.append(d_column_value)

#         listbox.delete(0, tk.END)
#         for item in listbox_items:
#             listbox.insert(tk.END, item)

# def on_listbox_select(event):
#     selection = event.widget.curselection()
#     if selection:
#         index = selection[0]
#         detail_text.delete('1.0', tk.END)
#         detail_text.insert(tk.INSERT, excel_data[index])

# app = tk.Tk()
# app.title("Excel Data Viewer")

# listbox_items = []
# excel_data = []

# load_button = tk.Button(app, text="Load Excel File", command=load_excel_data)
# load_button.pack(pady=10)

# listbox = Listbox(app, width=50, height=20)
# listbox.bind('<<ListboxSelect>>', on_listbox_select)
# listbox.pack(pady=10)

# detail_text = scrolledtext.ScrolledText(app, wrap=tk.WORD, width=50, height=10)
# detail_text.pack(pady=10)


import tkinter as tk
from tkinter import filedialog, Listbox, scrolledtext
import openpyxl
from tkinter.font import Font

def load_excel_data():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        listbox_items.clear()
        excel_data.clear()
        for row in sheet.iter_rows(min_row=2):  # Assuming the first row is the header
            e_column_value = row[4].value if row[4].value is not None else ""
            d_column_value = row[3].value if row[3].value is not None else ""
            listbox_items.append(e_column_value)
            excel_data.append(d_column_value)

        listbox.delete(0, tk.END)
        for item in listbox_items:
            listbox.insert(tk.END, item)

# def on_listbox_select(event):
#     selection = event.widget.curselection()
#     if selection:
#         index = selection[0]
#         detail_text.delete('1.0', tk.END)
#         detail_text.insert(tk.INSERT, excel_data[index])
        
def on_listbox_select(event):
    selection = event.widget.curselection()
    if selection:
        index = selection[0]
        detail_text.delete('1.0', tk.END)
        detail_data = excel_data[index]
        detail_text.insert(tk.INSERT, detail_data)
        
        # 클립보드에 데이터 복사
        app.clipboard_clear()  # 기존 클립보드 내용을 지웁니다
        app.clipboard_append(detail_data)  # 새로운 데이터를 클립보드에 추가합니다



app = tk.Tk()
app.title("Excel Data Viewer")

listbox_items = []
excel_data = []

load_button = tk.Button(app, text="Load Excel File", command=load_excel_data)
load_button.pack(pady=10)

listbox = Listbox(app, width=50, height=20)
listbox.bind('<<ListboxSelect>>', on_listbox_select)
listbox.pack(pady=10)

# 글자 크기를 더 크게 조정
text_font = Font(family="Helvetica", size=14)

detail_text = scrolledtext.ScrolledText(app, wrap=tk.WORD, width=60, height=15, font=text_font)
detail_text.pack(pady=10)

app.mainloop()
