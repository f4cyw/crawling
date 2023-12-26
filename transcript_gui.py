# # import tkinter as tk
# # from tkinter import scrolledtext
# # from youtube_transcript_api import YouTubeTranscriptApi
# # import re

# # def extract_video_id(youtube_url):
# #     match = re.search(r"v=([^&]+)", youtube_url) or re.search(r"youtu\.be/([^?]+)", youtube_url)
# #     return match.group(1) if match else None

# # def get_youtube_transcript(video_id):
# #     try:
# #         transcript = YouTubeTranscriptApi.get_transcript(video_id, languages=['ko','en'])
# #         transcript_text = '\n'.join([t['text'] for t in transcript])
# #         return transcript_text
# #     except Exception as e:
# #         return str(e)

# # def on_submit():
# #     youtube_url = url_entry.get()
# #     video_id = extract_video_id(youtube_url)
# #     transcript = get_youtube_transcript(video_id)
# #     result_text.delete('1.0', tk.END)
# #     result_text.insert(tk.INSERT, transcript)

# # app = tk.Tk()
# # app.title("YouTube Transcript Extractor")

# # tk.Label(app, text="YouTube Video URL:").pack()

# # url_entry = tk.Entry(app, width=50)
# # url_entry.pack()

# # submit_button = tk.Button(app, text="Get Transcript", command=on_submit)
# # submit_button.pack()

# # result_text = scrolledtext.ScrolledText(app, wrap=tk.WORD, width=100, height=100)
# # result_text.pack()

# # app.mainloop()


# import tkinter as tk
# from tkinter import filedialog
# from youtube_transcript_api import YouTubeTranscriptApi
# import re
# import openpyxl

# def extract_video_id(youtube_url):
#     match = re.search(r"v=([^&]+)", youtube_url) or re.search(r"youtu\.be/([^?]+)", youtube_url)
#     return match.group(1) if match else None

# def get_youtube_transcript(video_id):
#     try:
#         transcript = YouTubeTranscriptApi.get_transcript(video_id, languages=['ko', 'en'])
#         transcript_text = '\n'.join([t['text'] for t in transcript])
#         return transcript_text
#     except Exception as e:
#         return str(e)

# def read_links_from_excel(file_path):
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active
#     links = []
#     for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):  # Assuming links are in column B
#         if row[0].value:
#             links.append(row[0].value)
#     return links

# def save_to_excel(data, file_path):
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = 'YouTube Transcripts'
#     sheet.append(['Video URL', 'Transcript'])

#     for url, transcript in data:
#         sheet.append([url, transcript])

#     workbook.save(file_path)

# def process_and_save():
#     input_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
#     if input_file_path:
#         links = read_links_from_excel(input_file_path)
#         data = [(url, get_youtube_transcript(extract_video_id(url))) for url in links]

#         output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
#         if output_file_path:
#             save_to_excel(data, output_file_path)

# app = tk.Tk()
# app.title("YouTube Transcript Extractor")

# process_button = tk.Button(app, text="Process and Save Transcripts", command=process_and_save)
# process_button.pack(pady=20)

# app.mainloop()


import tkinter as tk
from tkinter import filedialog
from youtube_transcript_api import YouTubeTranscriptApi
import re
import openpyxl

def extract_video_id(youtube_url):
    match = re.search(r"v=([^&]+)", youtube_url) or re.search(r"youtu\.be/([^?]+)", youtube_url)
    return match.group(1) if match else None

def get_youtube_transcript(video_id):
    try:
        transcript = YouTubeTranscriptApi.get_transcript(video_id, languages=['ko', 'en'])
        transcript_text = '\n'.join([t['text'] for t in transcript])
        return transcript_text
    except Exception as e:
        return str(e)

def read_links_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    links = []
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):  # Assuming links are in column B
        if row[0].value:
            links.append(row[0].value)
    return links

def save_to_excel(data, file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'YouTube Transcripts'
    sheet.append(['Video URL', 'Transcript'])

    for url, transcript in data:
        sheet.append([url, transcript])

    workbook.save(file_path)

def update_progress_label(label, text):
    label.config(text=text)
    app.update()

def process_and_save():
    input_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if input_file_path:
        links = read_links_from_excel(input_file_path)
        data = []
        
        for i, url in enumerate(links, start=1):
            update_progress_label(progress_label, f"Processing {i}/{len(links)}: {url}")
            transcript = get_youtube_transcript(extract_video_id(url))
            data.append((url, transcript))

        output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if output_file_path:
            save_to_excel(data, output_file_path)
            update_progress_label(progress_label, "Completed")

app = tk.Tk()
app.title("YouTube Transcript Extractor")

process_button = tk.Button(app, text="Process and Save Transcripts", command=process_and_save)
process_button.pack(pady=10)

progress_label = tk.Label(app, text="")
progress_label.pack()

app.mainloop()
