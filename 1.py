import tkinter as tk
from tkinter import filedialog, Label, Frame, Canvas, Scrollbar, Button
from zipfile import ZipFile
from PIL import Image, ImageTk
import io
import pandas as pd
import datetime
import pdb
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as oimage
import openpyxl

# Global list to store image data
images_data = []
# Create a DataFrame to store image metadata
metadata_df = pd.DataFrame(columns=["ZIP Filename", "Image Filename", "Metadata", "positive", "negative", "steps"])
b=[]

# Define the MetadataViewer class
class MetadataViewer:
    def __init__(self, metadata_frame):
        self.metadata_frame = metadata_frame
        self.metadata_text = tk.Text(metadata_frame, wrap=tk.WORD)
        self.metadata_text.pack(side="top", fill="both", expand=True)

    def show_metadata(self, metadata):
        # Clear the current content
        self.metadata_text.delete('1.0', tk.END)

        if metadata:
            metadata_text = ""
            for key, value in metadata.items():
                readable_value = self.decode_metadata(value)
                formatted_value = MetadataViewer.format_multiline(readable_value)
                metadata_text += f"{key}: {formatted_value}\n"
            
            # Insert the metadata text into the Text widget
            self.metadata_text.insert(tk.END, metadata_text)
        else:
            self.metadata_text.insert(tk.END, "No metadata found")



    @staticmethod
    def decode_metadata(value):
        # Handle decoding of metadata if necessary (e.g., bytes to string)
        if isinstance(value, bytes):
            # Try to decode as UTF-8 after removing null bytes
            try:
                cleaned_value = value.replace(b'\x00', b'')
                # print(cleaned_value)
                return cleaned_value.decode('utf-8')
            except UnicodeDecodeError:
                # If UTF-8 decoding fails, try other common encodings
                encodings = ['ascii', 'windows-1250', 'windows-1252', 'iso-8859-1', 'iso-8859-15']
                for encoding in encodings:
                    try:
                        return cleaned_value.decode(encoding)
                    except UnicodeDecodeError:
                        continue
                # If all decodings fail, represent the binary data as hex
                return cleaned_value.hex()
        elif isinstance(value, str):
            return value
        else:
            return str(value)  # Convert non-string, non-bytes values to strings
    
    @staticmethod
    def format_multiline(text, max_line_length=40):
        lines = []  # List to hold each line
        current_line = ""  # String to hold the words for the current line

        for char in text:
            # Check if adding the character would exceed the max line length
            if len(current_line) >= max_line_length and char.isspace():
                lines.append(current_line)  # Add the current line to the lines list
                current_line = ""  # Start a new line
            else:
                current_line += char  # Add the character to the current line

        if current_line:
            lines.append(current_line)  # Add the last line if it's not empty

        return "\n".join(lines) # Join all lines with newline characters



def load_images_from_zip(filepath):
    images_data.clear()
    with ZipFile(filepath, 'r') as zip_ref:
        for file in zip_ref.namelist():
            if file.startswith("."):
                continue

            elif file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                # print(file)
                try:
                    image_data = zip_ref.read(file)
                    image = Image.open(io.BytesIO(image_data))
                    # print(file)
                    image.thumbnail((180, 180))
                    photo = ImageTk.PhotoImage(image)
                    images_data.append({"image":image, 'thumbnail': photo, 'metadata': image.info, 'filename': file})
                except Image.UnidentifiedImageError:
                    print(f"Cannot identify image file {file}")
    # print(image_data)
    return images_data

def thumbnail_click(event, image_info, metadata_viewer, zip_filename):
    
    metadata_viewer.show_metadata(image_info['metadata'])
    add_metadata_to_dataframe(zip_filename, image_info['filename'], image_info)

# Populate the inner frame with thumbnails
def populate_frame_with_thumbnails(images_data, inner_frame, metadata_viewer, selected_zip_filename):
    # Clear the frame first
    for widget in inner_frame.winfo_children():
        widget.destroy()
    
    # Populate the frame with image thumbnails and bind the click event
    for i, image_data in enumerate(images_data):
        thumb_label = tk.Label(inner_frame, image=image_data['thumbnail'])
        thumb_label.image = image_data['thumbnail']  # Keep a reference
        thumb_label.grid(row=i // 3, column=i % 3, padx=5, pady=5)
        # Pass selected_zip_filename to thumbnail_click
        thumb_label.bind('<Button-1>', lambda e, img_data=image_data, zip_filename=selected_zip_filename: thumbnail_click(e, img_data, metadata_viewer, zip_filename))

# Open the ZIP file, load the images, and populate the grid with thumbnails
def open_zip(inner_frame, metadata_viewer):
    filepath = filedialog.askopenfilename()
    if not filepath:
        return
    images_data = load_images_from_zip(filepath)  # This line was missing the assignment
    # Pass the selected ZIP filename to thumbnail_click
    for image_data in images_data:
        image_data['zip_filename'] = filepath
    populate_frame_with_thumbnails(images_data, inner_frame, metadata_viewer, filepath)  # Pass images_data and filepath here




def generate_filename():
    now = datetime.datetime.now()
    formatted_datetime = now.strftime("%Y%m%d%H%M")
    filename = f"{formatted_datetime}.xlsx"
    return filename


# Function to add metadata to the DataFrame
# Function to add metadata to the DataFrame
# Function to add metadata to the DataFrame
def add_metadata_to_dataframe(zip_filename, image_filename, image_info):
    global metadata_df
    metadata = metadata_viewer.metadata_text.get('1.0', tk.END)

    # Extract data for new columns J, K, M
    j2_data = extract_j2(metadata)
    k2_data = extract_k2(metadata)
    m2_data = extract_m2(metadata)

    # Extract data for positive, negative, and steps
    steps_data, remain_data = find_word_and_get_substring(metadata, "Steps:")
    negative_prompt, positive_prompt = find_word_and_get_substring(remain_data, "Negative prompt:")

    # Consolidate all data into a single dictionary
    new_data = {
        "ZIP Filename": zip_filename, 
        "Image Filename": image_filename, 
        "Metadata": metadata, 
        "positive": positive_prompt, 
        "negative": negative_prompt, 
        "steps": steps_data,
        "POSITIVE": j2_data, 
        "NEGATIVE": k2_data, 
        "SETTING": m2_data
    }

    # Add the consolidated data as a new row in the DataFrame
    metadata_df = pd.concat([metadata_df, pd.DataFrame([new_data])], ignore_index=True)
    image_info['image'].save(f"{image_filename}")


# Function to export the DataFrame to an Excel file
def export_to_excel():
    global metadata_df
    if not metadata_df.empty:
        print("Exporting data to Excel...")
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not filepath:
            print("User canceled the file dialog.")
            return  # User canceled the file dialog
        if not filepath.endswith(".xlsx"):
            filepath += ".xlsx"  # Ensure the file has the .xlsx extension

        # Specify UTF-8 encoding when saving to Excel
        metadata_df.to_excel(filepath, index=False, encoding="utf-8")
        
        image_in_excel(filepath)
        print(f"Data exported to {filepath}")


def find_word_and_get_substring(text, word):
    index = text.find(word)
    if index != -1:
        return text[index:], text[:index]
    else : 
        return "no data", "no data"
    


def image_in_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    a=[]
    # b=[]
    
    for i in range(ws.max_row-1):
        row_number=i+2
        column_letter = "J"
        cell_address = f'{column_letter}{row_number}'
        img_file = f'{ws.cell(row=i+2, column = 2).value}'
        # b.append(img_file)
        img = oimage(img_file)
        ws.add_image(img, cell_address)
    wb.save(filepath)
    delete_image_files_in_current_directory()
    # for x in b:
    #     os.remove(x)


    
# def process_all_zips():
#     for filename in os.listdir('.'):  # Assuming current directory, modify as needed
#         if filename.lower().endswith('.zip'):
#             print(f"Processing ZIP file: {filename}")
#             zip_filepath = os.path.join('.', filename)  # Modify as needed
#             try:
#                 images_data = load_images_from_zip(zip_filepath)
#                 for image_data in images_data:
#                     thumbnail_click(None, image_data, metadata_viewer, zip_filepath)
#             except:
#                 print(f"kkkkkkkkkkkkkkkkkkkkkk zip file: {filename}")
#                 b.append(filename)
#     print(b)
#     export_to_excel()     


def process_all_zips():
    total = len(os.listdir('.'))
    i=1
    for filename in os.listdir('.'):  # Assuming current directory, modify as needed
        if filename.lower().endswith('.zip') and not filename.lower().startswith("civitai"):
            print(f"{total} 중 {i} Processing ZIP file: {filename}")
            zip_filepath = os.path.join('.', filename)  # Modify as needed
            try:
                images_data = load_images_from_zip(zip_filepath)
                for image_data in images_data:
                    thumbnail_click(None, image_data, metadata_viewer, zip_filepath)
            except Exception as e:
                print(f"Error processing ZIP file {filename}: {e}")
                b.append(filename)
        i += 1
    # print(b)
    export_to_excel()



def delete_image_files_in_current_directory():
    # 현재 작업 디렉토리를 가져옵니다.
    current_directory = os.getcwd()

    # 현재 작업 디렉토리 내의 모든 파일 목록을 가져옵니다.
    files = os.listdir(current_directory)

    # 이미지 파일 확장자 리스트를 지정합니다. 여기에 원하는 이미지 확장자를 추가하세요.
    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp']

    # 현재 작업 디렉토리 내의 파일을 순회하며 이미지 파일인 경우 삭제합니다.
    for file in files:
        file_path = os.path.join(current_directory, file)
        # 파일이 존재하고 확장자가 이미지 확장자 중 하나인 경우 삭제합니다.
        if os.path.isfile(file_path) and any(file.lower().endswith(ext) for ext in image_extensions):
            os.remove(file_path)
            print(f"Deleted: {file_path}")


def preprocess_text(text):
    # Replace newlines with spaces and remove excess whitespace
    return ' '.join(text.split())

def extract_j2(text):
    text = preprocess_text(text)
    unicode_index = text.find("UNICODE")
    parameters_index = text.find("parameters:")
    
    # Determine the starting index based on which keyword appears first
    if unicode_index != -1 or parameters_index != -1:
        if unicode_index == -1 or (parameters_index != -1 and parameters_index < unicode_index):
            start_index = parameters_index + len("parameters:")
        else:
            start_index = unicode_index + len("UNICODE")

        end_index_negative = text.find("Negative prompt:", start_index)
        end_index_steps = text.find("Steps:", start_index)
        end_index = min(end_index_negative if end_index_negative != -1 else float('inf'),
                        end_index_steps if end_index_steps != -1 else float('inf'))
        if end_index != float('inf'):
            return text[start_index:end_index].strip()
    return ""



def extract_k2(text):
    text = preprocess_text(text)
    negative_index = text.find("Negative prompt:")
    if negative_index != -1:
        start_index = negative_index + 17  # Adjusted index for "Negative prompt:"
        end_index = text.find("Steps:", start_index)
        if end_index != -1:
            return text[start_index:end_index].strip()
    return ""

def extract_m2(text):
    text = preprocess_text(text)
    steps_index = text.find("Steps:")
    if steps_index != -1:
        # Start the substring from the index of "steps:" to include it
        return text[steps_index:].strip()
    return ""


# Main application setup
if __name__ == '__main__':
    root = tk.Tk()
    root.title("ZIP Image Metadata Viewer")
    root.geometry('1000x600')

    # Create the metadata display frame
    metadata_canvas = Canvas(root)
    metadata_frame = Frame(metadata_canvas)
    metadata_scrollbar = Scrollbar(root, orient='vertical', command=metadata_canvas.yview)
    metadata_canvas.configure(yscrollcommand=metadata_scrollbar.set)

    metadata_scrollbar.pack(side=tk.RIGHT, fill='y')
    metadata_canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

    metadata_canvas.create_window((0, 0), window=metadata_frame, anchor='nw')
    metadata_frame.bind('<Configure>', lambda e: metadata_canvas.configure(scrollregion=metadata_canvas.bbox("all")))

# Initialize the MetadataViewer
    metadata_viewer = MetadataViewer(metadata_frame)

    # Create the canvas and scrollbar for thumbnails
    canvas = Canvas(root)
    scrollbar = Scrollbar(root, orient='vertical', command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side=tk.RIGHT, fill='y')
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Inner frame for thumbnails
    inner_frame = Frame(canvas)
    canvas.create_window((0, 0), window=inner_frame, anchor='nw')
    inner_frame.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    # Initialize the MetadataViewer
   

    # Button to open a ZIP file
    open_button = Button(root, text="Open ZIP File", command=lambda: open_zip(inner_frame, metadata_viewer))
    open_button.pack(side=tk.TOP, fill='x')

    export_button = Button(root, text="Export to Excel", command=export_to_excel)
    export_button.pack(side=tk.TOP, fill='x')
    
    try:
        process_all_zips()

    except:
        delete_image_files_in_current_directory()
        print()
    print(b)

    root.mainloop()
